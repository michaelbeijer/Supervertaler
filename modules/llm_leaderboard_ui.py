"""
LLM Leaderboard - Qt UI Components
===================================

PyQt6 user interface for LLM translation benchmarking.

Features:
- Test dataset selection
- Model selection (checkboxes)
- Benchmark execution with progress
- Results table with comparison
- Summary statistics panel
- Export functionality

Author: Michael Beijer
License: MIT
"""

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGroupBox, QLabel,
    QPushButton, QComboBox, QCheckBox, QTableWidget, QTableWidgetItem,
    QProgressBar, QTextEdit, QSplitter, QHeaderView, QMessageBox,
    QFileDialog
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QColor, QFont
from typing import List, Optional, Dict
import json
from pathlib import Path

try:
    from modules.llm_leaderboard import (
        LLMLeaderboard, TestDataset, ModelConfig, BenchmarkResult,
        create_sample_datasets, CHRF_AVAILABLE
    )
except ImportError:
    from llm_leaderboard import (
        LLMLeaderboard, TestDataset, ModelConfig, BenchmarkResult,
        create_sample_datasets, CHRF_AVAILABLE
    )


class BenchmarkThread(QThread):
    """Background thread for running benchmarks without blocking UI"""

    progress_update = pyqtSignal(int, int, str)  # current, total, message
    finished = pyqtSignal(list)  # results
    error = pyqtSignal(str)  # error message

    def __init__(self, leaderboard: LLMLeaderboard, dataset: TestDataset, models: List[ModelConfig]):
        super().__init__()
        self.leaderboard = leaderboard
        self.dataset = dataset
        self.models = models

    def run(self):
        """Run benchmark in background thread"""
        try:
            results = self.leaderboard.run_benchmark(
                self.dataset,
                self.models,
                progress_callback=self._on_progress
            )
            self.finished.emit(results)
        except Exception as e:
            self.error.emit(str(e))

    def _on_progress(self, current: int, total: int, message: str):
        """Forward progress updates to main thread"""
        self.progress_update.emit(current, total, message)


class LLMLeaderboardUI(QWidget):
    """Main UI widget for LLM Leaderboard"""

    def __init__(self, parent=None, llm_client_factory=None):
        super().__init__(parent)
        self.parent_app = parent
        self.llm_client_factory = llm_client_factory
        self.leaderboard = None
        self.benchmark_thread = None
        self.current_results = []

        # Load sample datasets
        self.datasets = create_sample_datasets()
        self.current_dataset = self.datasets[0] if self.datasets else None

        self.init_ui()

    def init_ui(self):
        """Initialize the user interface"""
        layout = QVBoxLayout()
        layout.setSpacing(5)  # Tighter spacing for consistency
        layout.setContentsMargins(10, 10, 10, 10)

        # Header (matches TMX Editor / AutoFingers / PDF Rescue style)
        header = QLabel("🏆 LLM Leaderboard")
        header.setStyleSheet("font-size: 16pt; font-weight: bold; color: #1976D2;")
        layout.addWidget(header, 0)  # 0 = no stretch, stays compact

        # Description box (matches TMX Editor / AutoFingers / PDF Rescue style)
        description = QLabel(
            "Translation Quality Benchmarking System - A Supervertaler Module.\n"
            "Compare translation quality, speed, and cost across multiple LLM providers."
        )
        description.setWordWrap(True)
        description.setStyleSheet("color: #666; padding: 5px; background-color: #E3F2FD; border-radius: 3px;")
        layout.addWidget(description, 0)

        # Spacing after description
        layout.addSpacing(10)

        # Top section: Dataset and Model selection
        top_widget = self._create_top_section()
        layout.addWidget(top_widget)

        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)

        # Status label
        self.status_label = QLabel("Ready")
        self.status_label.setStyleSheet("color: #666; font-size: 9pt;")
        layout.addWidget(self.status_label)

        # Splitter for results and log
        splitter = QSplitter(Qt.Orientation.Vertical)

        # Results table
        self.results_table = self._create_results_table()
        splitter.addWidget(self.results_table)

        # Summary panel
        self.summary_panel = self._create_summary_panel()
        splitter.addWidget(self.summary_panel)

        # Log output
        self.log_output = QTextEdit()
        self.log_output.setReadOnly(True)
        self.log_output.setMaximumHeight(150)
        self.log_output.setPlaceholderText("Benchmark log will appear here...")
        splitter.addWidget(self.log_output)

        splitter.setStretchFactor(0, 3)  # Results table gets most space
        splitter.setStretchFactor(1, 1)  # Summary panel medium space
        splitter.setStretchFactor(2, 1)  # Log output smallest

        layout.addWidget(splitter)

        self.setLayout(layout)

    def _create_top_section(self) -> QWidget:
        """Create dataset selection and model selection section"""
        widget = QWidget()
        layout = QHBoxLayout()

        # Left: Dataset selection
        dataset_group = QGroupBox("Test Dataset")
        dataset_layout = QVBoxLayout()

        self.dataset_combo = QComboBox()
        for ds in self.datasets:
            self.dataset_combo.addItem(f"{ds.name} ({len(ds.segments)} segments)", ds)
        self.dataset_combo.currentIndexChanged.connect(self._on_dataset_changed)
        dataset_layout.addWidget(self.dataset_combo)

        dataset_info_label = QLabel("Select a test set to compare models")
        dataset_info_label.setStyleSheet("color: #666; font-size: 9pt;")
        dataset_layout.addWidget(dataset_info_label)

        # TODO: Add Import CSV and Create Custom buttons
        # import_btn = QPushButton("Import CSV...")
        # dataset_layout.addWidget(import_btn)

        dataset_layout.addStretch()
        dataset_group.setLayout(dataset_layout)
        layout.addWidget(dataset_group)

        # Right: Model selection
        model_group = QGroupBox("Model Selection")
        model_layout = QVBoxLayout()

        model_layout.addWidget(QLabel("Select models to test:"))

        # OpenAI models
        self.openai_checkbox = QCheckBox("OpenAI (GPT-4o)")
        self.openai_checkbox.setChecked(True)
        model_layout.addWidget(self.openai_checkbox)

        self.openai_model_combo = QComboBox()
        self.openai_model_combo.addItems([
            "gpt-4o",
            "gpt-4o-mini",
            "gpt-5"
        ])
        self.openai_model_combo.setEnabled(True)
        model_layout.addWidget(self.openai_model_combo)

        # Claude models
        self.claude_checkbox = QCheckBox("Claude (Sonnet 4.5)")
        self.claude_checkbox.setChecked(True)
        model_layout.addWidget(self.claude_checkbox)

        self.claude_model_combo = QComboBox()
        self.claude_model_combo.addItems([
            "claude-sonnet-4-5-20250929",
            "claude-haiku-4-5-20251001",
            "claude-opus-4-1-20250805"
        ])
        self.claude_model_combo.setEnabled(True)
        model_layout.addWidget(self.claude_model_combo)

        # Gemini models
        self.gemini_checkbox = QCheckBox("Gemini (2.5 Flash)")
        self.gemini_checkbox.setChecked(True)
        model_layout.addWidget(self.gemini_checkbox)

        self.gemini_model_combo = QComboBox()
        self.gemini_model_combo.addItems([
            "gemini-2.5-flash",
            "gemini-2.5-flash-lite",
            "gemini-2.5-pro"
        ])
        self.gemini_model_combo.setEnabled(True)
        model_layout.addWidget(self.gemini_model_combo)

        model_layout.addStretch()

        # Run button
        self.run_button = QPushButton("🚀 Run Benchmark")
        self.run_button.setStyleSheet("font-weight: bold; padding: 8px;")
        self.run_button.clicked.connect(self._on_run_benchmark)
        model_layout.addWidget(self.run_button)

        # Cancel button
        self.cancel_button = QPushButton("Cancel")
        self.cancel_button.setEnabled(False)
        self.cancel_button.clicked.connect(self._on_cancel_benchmark)
        model_layout.addWidget(self.cancel_button)

        # Export button
        self.export_button = QPushButton("📊 Export Results...")
        self.export_button.setEnabled(False)
        self.export_button.clicked.connect(self._on_export_results)
        model_layout.addWidget(self.export_button)

        model_group.setLayout(model_layout)
        layout.addWidget(model_group)

        widget.setLayout(layout)
        return widget

    def _create_results_table(self) -> QTableWidget:
        """Create results comparison table"""
        table = QTableWidget()
        table.setColumnCount(6)
        table.setHorizontalHeaderLabels([
            "Segment", "Source Text", "Model", "Translation", "Speed (ms)", "Quality"
        ])

        # Set column widths
        header = table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)

        table.setAlternatingRowColors(True)
        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)

        return table

    def _create_summary_panel(self) -> QWidget:
        """Create summary statistics panel"""
        widget = QGroupBox("Summary Statistics")
        layout = QVBoxLayout()

        self.summary_table = QTableWidget()
        self.summary_table.setColumnCount(5)
        self.summary_table.setHorizontalHeaderLabels([
            "Model", "Avg Speed (ms)", "Avg Quality", "Success", "Errors"
        ])

        header = self.summary_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

        self.summary_table.setMaximumHeight(200)
        self.summary_table.setAlternatingRowColors(True)
        self.summary_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

        layout.addWidget(self.summary_table)
        widget.setLayout(layout)

        return widget

    def _on_dataset_changed(self, index: int):
        """Handle dataset selection change"""
        self.current_dataset = self.dataset_combo.itemData(index)
        self.log(f"Selected dataset: {self.current_dataset.name}")

    def _on_run_benchmark(self):
        """Start benchmark execution"""
        if not self.llm_client_factory:
            QMessageBox.warning(self, "Error", "LLM client factory not available")
            return

        if not self.current_dataset:
            QMessageBox.warning(self, "Error", "No dataset selected")
            return

        # Get selected models
        models = self._get_selected_models()
        if not models:
            QMessageBox.warning(self, "Error", "Please select at least one model to test")
            return

        # Confirm if sacrebleu not available
        if not CHRF_AVAILABLE:
            reply = QMessageBox.question(
                self,
                "Quality Scoring Unavailable",
                "sacrebleu library is not installed. Quality scores will not be calculated.\n\n"
                "Continue anyway?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.No:
                return

        # Clear previous results
        self.results_table.setRowCount(0)
        self.summary_table.setRowCount(0)
        self.log_output.clear()
        self.current_results = []

        # Update UI state
        self.run_button.setEnabled(False)
        self.cancel_button.setEnabled(True)
        self.export_button.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.status_label.setText("Running benchmark...")

        # Create leaderboard instance
        self.leaderboard = LLMLeaderboard(self.llm_client_factory, self.log)

        # Start benchmark in background thread
        self.benchmark_thread = BenchmarkThread(self.leaderboard, self.current_dataset, models)
        self.benchmark_thread.progress_update.connect(self._on_progress_update)
        self.benchmark_thread.finished.connect(self._on_benchmark_finished)
        self.benchmark_thread.error.connect(self._on_benchmark_error)
        self.benchmark_thread.start()

    def _on_cancel_benchmark(self):
        """Cancel running benchmark"""
        if self.leaderboard:
            self.leaderboard.cancel_benchmark()
            self.log("⚠️ Cancelling benchmark...")

    def _on_progress_update(self, current: int, total: int, message: str):
        """Update progress bar and status"""
        self.progress_bar.setMaximum(total)
        self.progress_bar.setValue(current)
        self.status_label.setText(f"{message} ({current}/{total})")

    def _on_benchmark_finished(self, results: List[BenchmarkResult]):
        """Handle benchmark completion"""
        self.current_results = results

        # Update UI state
        self.run_button.setEnabled(True)
        self.cancel_button.setEnabled(False)
        self.export_button.setEnabled(True)
        self.progress_bar.setVisible(False)
        self.status_label.setText(f"✅ Benchmark complete: {len(results)} results")

        # Populate results table
        self._populate_results_table(results)

        # Populate summary table
        self._populate_summary_table()

        self.log("✅ Benchmark finished successfully")

    def _on_benchmark_error(self, error_msg: str):
        """Handle benchmark error"""
        self.run_button.setEnabled(True)
        self.cancel_button.setEnabled(False)
        self.progress_bar.setVisible(False)
        self.status_label.setText("❌ Benchmark failed")

        QMessageBox.critical(self, "Benchmark Error", f"An error occurred:\n\n{error_msg}")
        self.log(f"❌ Error: {error_msg}")

    def _populate_results_table(self, results: List[BenchmarkResult]):
        """Populate results table with benchmark data"""
        # Group results by segment
        segments_dict = {}
        for result in results:
            if result.segment_id not in segments_dict:
                segments_dict[result.segment_id] = []
            segments_dict[result.segment_id].append(result)

        # Populate table
        row = 0
        for segment_id in sorted(segments_dict.keys()):
            segment_results = segments_dict[segment_id]

            # Get source text from dataset
            source_text = ""
            for seg in self.current_dataset.segments:
                if seg.id == segment_id:
                    source_text = seg.source
                    break

            # Truncate source text for display
            if len(source_text) > 80:
                source_text = source_text[:77] + "..."

            for result in segment_results:
                self.results_table.insertRow(row)

                # Segment ID
                self.results_table.setItem(row, 0, QTableWidgetItem(str(segment_id)))

                # Source text
                self.results_table.setItem(row, 1, QTableWidgetItem(source_text))

                # Model name
                self.results_table.setItem(row, 2, QTableWidgetItem(result.model_name))

                # Translation output
                output_text = result.output if result.output else f"ERROR: {result.error}"
                if len(output_text) > 100:
                    output_text = output_text[:97] + "..."
                item = QTableWidgetItem(output_text)
                if result.error:
                    item.setForeground(QColor("red"))
                self.results_table.setItem(row, 3, item)

                # Speed
                speed_item = QTableWidgetItem(f"{result.latency_ms:.0f}")
                self.results_table.setItem(row, 4, speed_item)

                # Quality
                if result.quality_score is not None:
                    quality_item = QTableWidgetItem(f"{result.quality_score:.1f}")
                    self.results_table.setItem(row, 5, quality_item)
                else:
                    self.results_table.setItem(row, 5, QTableWidgetItem("—"))

                row += 1

    def _populate_summary_table(self):
        """Populate summary statistics table"""
        if not self.leaderboard:
            return

        summary = self.leaderboard.get_summary_stats()

        self.summary_table.setRowCount(len(summary))
        row = 0

        for model_name, stats in summary.items():
            # Model name
            self.summary_table.setItem(row, 0, QTableWidgetItem(model_name))

            # Avg speed
            avg_speed = stats["avg_latency_ms"]
            speed_item = QTableWidgetItem(f"{avg_speed:.0f}")
            self.summary_table.setItem(row, 1, speed_item)

            # Avg quality
            avg_quality = stats["avg_quality_score"]
            if avg_quality is not None:
                quality_item = QTableWidgetItem(f"{avg_quality:.1f}")
                self.summary_table.setItem(row, 2, quality_item)
            else:
                self.summary_table.setItem(row, 2, QTableWidgetItem("—"))

            # Success count
            self.summary_table.setItem(row, 3, QTableWidgetItem(str(stats["success_count"])))

            # Error count
            error_item = QTableWidgetItem(str(stats["error_count"]))
            if stats["error_count"] > 0:
                error_item.setForeground(QColor("red"))
            self.summary_table.setItem(row, 4, error_item)

            row += 1

    def _on_export_results(self):
        """Export results to file (JSON or Excel)"""
        if not self.current_results:
            QMessageBox.warning(self, "No Results", "No benchmark results to export")
            return

        # Generate filename with dataset info
        dataset_name = self.current_dataset.name.replace(" ", "_").replace("→", "-")
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"LLM_Leaderboard_{dataset_name}_{timestamp}.xlsx"

        # Ask user for file path and format
        filepath, selected_filter = QFileDialog.getSaveFileName(
            self,
            "Export Benchmark Results",
            default_filename,
            "Excel Files (*.xlsx);;JSON Files (*.json);;All Files (*)"
        )

        if not filepath:
            return

        try:
            # Determine export format from selected filter or file extension
            if "Excel" in selected_filter or filepath.endswith('.xlsx'):
                self._export_to_excel(filepath)
            else:
                self._export_to_json(filepath)

            QMessageBox.information(self, "Export Complete", f"Results exported to:\n{filepath}")
            self.log(f"OK Results exported to {filepath}")

        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export results:\n{str(e)}")
            self.log(f"ERROR Export error: {e}")

    def _export_to_json(self, filepath: str):
        """Export results to JSON file"""
        export_data = self.leaderboard.export_to_dict()
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False)

    def _export_to_excel(self, filepath: str):
        """Export results to Excel file with title sheet, detailed results, and summary"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime

        wb = Workbook()

        # === TITLE/INFO SHEET ===
        ws_info = wb.active
        ws_info.title = "About"

        # Title with emoji (matches UI header style)
        ws_info['A1'] = "🏆 LLM Leaderboard"
        ws_info['A1'].font = Font(size=24, bold=True, color="1976D2")  # Blue color matching UI
        ws_info.merge_cells('A1:D1')

        # Subtitle (matches UI description style)
        ws_info['A2'] = "Translation Quality Benchmarking System"
        ws_info['A2'].font = Font(size=12, italic=True, color="666666")
        ws_info.merge_cells('A2:D2')

        # Supervertaler module branding (matches standardized naming)
        ws_info['A3'] = "A Supervertaler Module"
        ws_info['A3'].font = Font(size=11, color="0066CC", underline="single")
        ws_info['A3'].hyperlink = "https://supervertaler.com/"
        ws_info.merge_cells('A3:D3')

        # Spacing
        ws_info.row_dimensions[4].height = 20

        # Benchmark Info
        info_header_font = Font(bold=True, size=11)
        info_label_font = Font(size=10, color="666666")
        info_value_font = Font(size=10)

        row = 5
        ws_info[f'A{row}'] = "BENCHMARK INFORMATION"
        ws_info[f'A{row}'].font = Font(bold=True, size=12, color="366092")
        row += 1

        # Dataset info
        ws_info[f'A{row}'] = "Test Dataset:"
        ws_info[f'A{row}'].font = info_label_font
        ws_info[f'B{row}'] = self.current_dataset.name
        ws_info[f'B{row}'].font = info_value_font
        row += 1

        ws_info[f'A{row}'] = "Description:"
        ws_info[f'A{row}'].font = info_label_font
        ws_info[f'B{row}'] = self.current_dataset.description
        ws_info[f'B{row}'].font = info_value_font
        row += 1

        ws_info[f'A{row}'] = "Segments Tested:"
        ws_info[f'A{row}'].font = info_label_font
        ws_info[f'B{row}'] = len(self.current_dataset.segments)
        ws_info[f'B{row}'].font = info_value_font
        row += 1

        ws_info[f'A{row}'] = "Date & Time:"
        ws_info[f'A{row}'].font = info_label_font
        ws_info[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_info[f'B{row}'].font = info_value_font
        row += 1

        # Models tested
        ws_info[f'A{row}'] = "Models Tested:"
        ws_info[f'A{row}'].font = info_label_font
        models_tested = set(r.model_name for r in self.current_results)
        ws_info[f'B{row}'] = ", ".join(models_tested)
        ws_info[f'B{row}'].font = info_value_font
        row += 2

        # Explanation section
        ws_info[f'A{row}'] = "UNDERSTANDING THE RESULTS"
        ws_info[f'A{row}'].font = Font(bold=True, size=12, color="366092")
        row += 1

        explanations = [
            ("Quality Score (chrF++):", "Character-level metric measuring translation accuracy. Higher is better (0-100). Scores above 80 indicate excellent quality."),
            ("Speed (ms):", "Translation time in milliseconds. Lower is better. Typical range: 1000-5000ms per segment."),
            ("Success Count:", "Number of segments successfully translated without errors."),
            ("Error Count:", "Number of failed translations. Should be 0 for production use."),
            ("Color Coding:", "Green highlighting indicates the best performer in each category (quality/speed).")
        ]

        for label, explanation in explanations:
            ws_info[f'A{row}'] = label
            ws_info[f'A{row}'].font = Font(bold=True, size=10)
            ws_info[f'B{row}'] = explanation
            ws_info[f'B{row}'].font = Font(size=10)
            ws_info[f'B{row}'].alignment = Alignment(wrap_text=True)
            row += 1

        row += 1
        ws_info[f'A{row}'] = "NAVIGATION"
        ws_info[f'A{row}'].font = Font(bold=True, size=12, color="366092")
        row += 1

        navigation_items = [
            ("Summary Tab:", "View aggregated statistics and compare models side-by-side."),
            ("Results Tab:", "View detailed translation output for each segment and model."),
        ]

        for label, description in navigation_items:
            ws_info[f'A{row}'] = label
            ws_info[f'A{row}'].font = Font(bold=True, size=10)
            ws_info[f'B{row}'] = description
            ws_info[f'B{row}'].font = Font(size=10)
            row += 1

        # Column widths
        ws_info.column_dimensions['A'].width = 20
        ws_info.column_dimensions['B'].width = 80
        ws_info.column_dimensions['C'].width = 15
        ws_info.column_dimensions['D'].width = 15

        # === SUMMARY SHEET ===
        ws_summary = wb.create_sheet("Summary")

        # Header row
        summary_headers = ["Model", "Provider", "Model ID", "Avg Speed (ms)", "Avg Quality (chrF++)",
                          "Success Count", "Error Count", "Total Tests"]
        ws_summary.append(summary_headers)

        # Format header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col_num, _ in enumerate(summary_headers, 1):
            cell = ws_summary.cell(1, col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Get summary statistics
        summary_stats = self.leaderboard.get_summary_stats()

        # Populate summary data
        row_num = 2
        for model_name, stats in summary_stats.items():
            # Extract provider and model ID from results
            provider = ""
            model_id = ""
            for result in self.current_results:
                if result.model_name == model_name:
                    provider = result.provider
                    model_id = result.model_id
                    break

            total_tests = stats["success_count"] + stats["error_count"]

            ws_summary.cell(row_num, 1, model_name)
            ws_summary.cell(row_num, 2, provider)
            ws_summary.cell(row_num, 3, model_id)
            ws_summary.cell(row_num, 4, f"{stats['avg_latency_ms']:.0f}" if stats['avg_latency_ms'] else "")
            ws_summary.cell(row_num, 5, f"{stats['avg_quality_score']:.2f}" if stats['avg_quality_score'] else "")
            ws_summary.cell(row_num, 6, stats["success_count"])
            ws_summary.cell(row_num, 7, stats["error_count"])
            ws_summary.cell(row_num, 8, total_tests)

            # Highlight best quality score
            if stats['avg_quality_score']:
                quality_cell = ws_summary.cell(row_num, 5)
                best_quality = max(s['avg_quality_score'] for s in summary_stats.values() if s['avg_quality_score'])
                if stats['avg_quality_score'] == best_quality:
                    quality_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    quality_cell.font = Font(bold=True)

            # Highlight best speed
            if stats['avg_latency_ms']:
                speed_cell = ws_summary.cell(row_num, 4)
                best_speed = min(s['avg_latency_ms'] for s in summary_stats.values() if s['avg_latency_ms'])
                if stats['avg_latency_ms'] == best_speed:
                    speed_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    speed_cell.font = Font(bold=True)

            row_num += 1

        # Auto-size summary columns
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 15
        ws_summary.column_dimensions['C'].width = 30
        ws_summary.column_dimensions['D'].width = 18
        ws_summary.column_dimensions['E'].width = 20
        ws_summary.column_dimensions['F'].width = 15
        ws_summary.column_dimensions['G'].width = 15
        ws_summary.column_dimensions['H'].width = 15

        # === RESULTS SHEET ===
        ws_results = wb.create_sheet("Results")

        # Header row
        headers = ["Segment ID", "Source Text", "Reference", "Model", "Provider",
                   "Translation Output", "Speed (ms)", "Quality (chrF++)", "Error"]
        ws_results.append(headers)

        # Format header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col_num, _ in enumerate(headers, 1):
            cell = ws_results.cell(1, col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Group results by segment
        segments_dict = {}
        for result in self.current_results:
            if result.segment_id not in segments_dict:
                segments_dict[result.segment_id] = []
            segments_dict[result.segment_id].append(result)

        # Populate data rows
        row_num = 2
        for segment_id in sorted(segments_dict.keys()):
            segment_results = segments_dict[segment_id]

            # Get source and reference text from dataset
            source_text = ""
            reference_text = ""
            for seg in self.current_dataset.segments:
                if seg.id == segment_id:
                    source_text = seg.source
                    reference_text = seg.reference
                    break

            for result in segment_results:
                ws_results.cell(row_num, 1, segment_id)
                ws_results.cell(row_num, 2, source_text)
                ws_results.cell(row_num, 3, reference_text)
                ws_results.cell(row_num, 4, result.model_name)
                ws_results.cell(row_num, 5, result.provider)
                ws_results.cell(row_num, 6, result.output if result.output else "")
                ws_results.cell(row_num, 7, f"{result.latency_ms:.0f}" if result.latency_ms else "")
                ws_results.cell(row_num, 8, f"{result.quality_score:.2f}" if result.quality_score else "")
                ws_results.cell(row_num, 9, result.error if result.error else "")

                # Highlight errors in red
                if result.error:
                    error_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    for col in range(1, 10):
                        ws_results.cell(row_num, col).fill = error_fill

                row_num += 1

        # Auto-size columns
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for cell in ws_results[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)  # Cap at 80 for readability
            ws_results.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        wb.save(filepath)

    def _get_selected_models(self) -> List[ModelConfig]:
        """Get list of selected models from UI"""
        models = []

        # Map model IDs to friendly display names
        model_names = {
            # OpenAI
            "gpt-4o": "GPT-4o",
            "gpt-4o-mini": "GPT-4o Mini",
            "gpt-5": "GPT-5 (Reasoning)",

            # Claude
            "claude-sonnet-4-5-20250929": "Claude Sonnet 4.5",
            "claude-haiku-4-5-20251001": "Claude Haiku 4.5",
            "claude-opus-4-1-20250805": "Claude Opus 4.1",

            # Gemini
            "gemini-2.5-flash": "Gemini 2.5 Flash",
            "gemini-2.5-flash-lite": "Gemini 2.5 Flash Lite",
            "gemini-2.5-pro": "Gemini 2.5 Pro",
            "gemini-2.0-flash-exp": "Gemini 2.0 Flash (Exp)"
        }

        if self.openai_checkbox.isChecked():
            model_id = self.openai_model_combo.currentText()
            models.append(ModelConfig(
                name=model_names.get(model_id, model_id),
                provider="openai",
                model_id=model_id
            ))

        if self.claude_checkbox.isChecked():
            model_id = self.claude_model_combo.currentText()
            models.append(ModelConfig(
                name=model_names.get(model_id, model_id),
                provider="claude",
                model_id=model_id
            ))

        if self.gemini_checkbox.isChecked():
            model_id = self.gemini_model_combo.currentText()
            models.append(ModelConfig(
                name=model_names.get(model_id, model_id),
                provider="gemini",
                model_id=model_id
            ))

        return models

    def log(self, message: str):
        """Append message to log output and auto-scroll to bottom"""
        self.log_output.append(message)
        # Auto-scroll to bottom to show latest messages
        scrollbar = self.log_output.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())


# For standalone testing
if __name__ == "__main__":
    import sys
    from PyQt6.QtWidgets import QApplication

    app = QApplication(sys.argv)

    # Mock LLM client factory for testing UI
    def mock_llm_factory(provider, model):
        print(f"Mock: Creating {provider} client with model {model}")
        return None

    window = LLMLeaderboardUI(llm_client_factory=mock_llm_factory)
    window.setWindowTitle("LLM Leaderboard")
    window.resize(1200, 800)
    window.show()

    sys.exit(app.exec())
